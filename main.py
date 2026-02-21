import os
import time
import json
import sqlite3
import threading
from collections import Counter
from typing import Dict, List, Tuple, Any, Optional

import requests
import telebot
import telebot.apihelper as apihelper
from telebot import types
from openpyxl import load_workbook
from apscheduler.schedulers.background import BackgroundScheduler
from apscheduler.triggers.cron import CronTrigger
from flask import Flask

# ================== НАСТРОЙКИ ==================
BOT_TOKEN = os.environ.get("BOT_TOKEN")
if not BOT_TOKEN:
    raise RuntimeError("BOT_TOKEN is missing in environment variables")

# Чтобы /export работал только у тебя:
# На Render добавь env var: ADMIN_CHAT_ID = 2106404214 (твой chat_id)
ADMIN_CHAT_ID = int(os.environ.get("ADMIN_CHAT_ID", "0"))

DEFAULT_TZ = "Europe/Berlin"
SEP = "||"
HISTORY_LIMIT = 60

PORT_DEFAULT = "10000"
DB_PATH = os.environ.get("DB_PATH", "bot.db")

# Telegram timeouts (меньше отвалов)
apihelper.CONNECT_TIMEOUT = 10
apihelper.READ_TIMEOUT = 30

bot = telebot.TeleBot(BOT_TOKEN, threaded=True, num_threads=4)

# ================== safe_send ==================
def safe_send(chat_id: int, text: str, reply_markup=None, tries: int = 3):
    for i in range(tries):
        try:
            return bot.send_message(chat_id, text, reply_markup=reply_markup)
        except (requests.exceptions.RequestException, ConnectionError):
            time.sleep(2 + i * 2)
        except Exception:
            time.sleep(1)
    return None

# ================== Flask (порт-заглушка для Render Web Service) ==================
app = Flask(__name__)

@app.get("/")
def home():
    return "OK", 200

# ================== Scheduler ==================
scheduler = BackgroundScheduler(timezone=DEFAULT_TZ)
scheduler.start()
scheduled_jobs: Dict[int, str] = {}

# ================== SQLite ==================
db_lock = threading.Lock()

def db_conn() -> sqlite3.Connection:
    con = sqlite3.connect(DB_PATH, check_same_thread=False)
    con.row_factory = sqlite3.Row
    # важная штука для каскадного удаления
    con.execute("PRAGMA foreign_keys = ON;")
    return con

CON = db_conn()

def init_db():
    with db_lock:
        cur = CON.cursor()
        cur.execute("""
        CREATE TABLE IF NOT EXISTS users (
            chat_id INTEGER PRIMARY KEY,
            reminder_enabled INTEGER NOT NULL DEFAULT 0,
            reminder_time TEXT DEFAULT NULL,
            awaiting_time INTEGER NOT NULL DEFAULT 0,
            last_overall REAL DEFAULT NULL,
            last_averages_json TEXT DEFAULT NULL
        )
        """)
        cur.execute("""
        CREATE TABLE IF NOT EXISTS grade_counter (
            chat_id INTEGER NOT NULL,
            k TEXT NOT NULL,
            cnt INTEGER NOT NULL,
            PRIMARY KEY (chat_id, k),
            FOREIGN KEY (chat_id) REFERENCES users(chat_id) ON DELETE CASCADE
        )
        """)
        cur.execute("""
        CREATE TABLE IF NOT EXISTS snapshots (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            chat_id INTEGER NOT NULL,
            ts TEXT NOT NULL,
            overall REAL NOT NULL,
            averages_json TEXT NOT NULL,
            FOREIGN KEY (chat_id) REFERENCES users(chat_id) ON DELETE CASCADE
        )
        """)
        # для undo: хранение counter на каждый snapshot
        cur.execute("""
        CREATE TABLE IF NOT EXISTS counter_snapshots (
            snapshot_id INTEGER NOT NULL,
            k TEXT NOT NULL,
            cnt INTEGER NOT NULL,
            PRIMARY KEY (snapshot_id, k),
            FOREIGN KEY (snapshot_id) REFERENCES snapshots(id) ON DELETE CASCADE
        )
        """)
        CON.commit()

def ensure_user(chat_id: int):
    with db_lock:
        cur = CON.cursor()
        cur.execute("INSERT OR IGNORE INTO users(chat_id) VALUES (?)", (chat_id,))
        CON.commit()

def get_user_row(chat_id: int) -> sqlite3.Row:
    ensure_user(chat_id)
    with db_lock:
        cur = CON.cursor()
        cur.execute("SELECT * FROM users WHERE chat_id=?", (chat_id,))
        return cur.fetchone()

def set_user_fields(chat_id: int, **fields):
    ensure_user(chat_id)
    if not fields:
        return
    cols, vals = [], []
    for k, v in fields.items():
        cols.append(f"{k}=?")
        vals.append(v)
    vals.append(chat_id)
    q = "UPDATE users SET " + ", ".join(cols) + " WHERE chat_id=?"
    with db_lock:
        cur = CON.cursor()
        cur.execute(q, tuple(vals))
        CON.commit()

def get_counter(chat_id: int) -> Counter:
    ensure_user(chat_id)
    with db_lock:
        cur = CON.cursor()
        cur.execute("SELECT k, cnt FROM grade_counter WHERE chat_id=?", (chat_id,))
        rows = cur.fetchall()
    c = Counter()
    for r in rows:
        c[r["k"]] = int(r["cnt"])
    return c

def set_counter(chat_id: int, c: Counter):
    ensure_user(chat_id)
    with db_lock:
        cur = CON.cursor()
        cur.execute("DELETE FROM grade_counter WHERE chat_id=?", (chat_id,))
        cur.executemany(
            "INSERT INTO grade_counter(chat_id, k, cnt) VALUES (?, ?, ?)",
            [(chat_id, k, int(v)) for k, v in c.items()]
        )
        CON.commit()

def add_snapshot(chat_id: int, ts: str, overall: float, averages: Dict[str, float]) -> int:
    ensure_user(chat_id)
    averages_json = json.dumps(averages, ensure_ascii=False)
    with db_lock:
        cur = CON.cursor()
        cur.execute(
            "INSERT INTO snapshots(chat_id, ts, overall, averages_json) VALUES (?, ?, ?, ?)",
            (chat_id, ts, float(overall), averages_json)
        )
        snapshot_id = int(cur.lastrowid)

        # ограничим историю до HISTORY_LIMIT
        cur.execute("""
            DELETE FROM snapshots
            WHERE chat_id=?
              AND id NOT IN (
                SELECT id FROM snapshots WHERE chat_id=? ORDER BY id DESC LIMIT ?
              )
        """, (chat_id, chat_id, HISTORY_LIMIT))

        CON.commit()
    return snapshot_id

def save_counter_snapshot(snapshot_id: int, c: Counter):
    with db_lock:
        cur = CON.cursor()
        cur.executemany(
            "INSERT OR REPLACE INTO counter_snapshots(snapshot_id, k, cnt) VALUES (?, ?, ?)",
            [(snapshot_id, k, int(v)) for k, v in c.items()]
        )
        CON.commit()

def get_latest_snapshot_id(chat_id: int) -> Optional[int]:
    with db_lock:
        cur = CON.cursor()
        cur.execute("SELECT id FROM snapshots WHERE chat_id=? ORDER BY id DESC LIMIT 1", (chat_id,))
        row = cur.fetchone()
    return int(row["id"]) if row else None

def get_snapshot_data(snapshot_id: int) -> Optional[Dict[str, Any]]:
    with db_lock:
        cur = CON.cursor()
        cur.execute("SELECT ts, overall, averages_json FROM snapshots WHERE id=?", (snapshot_id,))
        row = cur.fetchone()
    if not row:
        return None
    return {
        "ts": row["ts"],
        "overall": float(row["overall"]),
        "averages": json.loads(row["averages_json"]),
    }

def get_counter_by_snapshot(snapshot_id: int) -> Counter:
    with db_lock:
        cur = CON.cursor()
        cur.execute("SELECT k, cnt FROM counter_snapshots WHERE snapshot_id=?", (snapshot_id,))
        rows = cur.fetchall()
    c = Counter()
    for r in rows:
        c[r["k"]] = int(r["cnt"])
    return c

def get_history(chat_id: int, limit: int = 10) -> List[Dict[str, Any]]:
    ensure_user(chat_id)
    with db_lock:
        cur = CON.cursor()
        cur.execute("""
            SELECT ts, overall, averages_json
            FROM snapshots
            WHERE chat_id=?
            ORDER BY id DESC
            LIMIT ?
        """, (chat_id, limit))
        rows = cur.fetchall()
    out = []
    for r in reversed(rows):
        out.append({
            "ts": r["ts"],
            "overall": float(r["overall"]),
            "averages": json.loads(r["averages_json"])
        })
    return out

def get_last_averages(chat_id: int) -> Dict[str, float]:
    row = get_user_row(chat_id)
    js = row["last_averages_json"]
    if not js:
        return {}
    try:
        return json.loads(js)
    except Exception:
        return {}

def get_last_overall(chat_id: int) -> Optional[float]:
    row = get_user_row(chat_id)
    v = row["last_overall"]
    return None if v is None else float(v)

init_db()

# ================== Export (все пользователи) ==================
def dump_all_data_to_dict() -> Dict[str, Any]:
    with db_lock:
        cur = CON.cursor()

        cur.execute("SELECT * FROM users")
        users = [dict(r) for r in cur.fetchall()]

        cur.execute("SELECT * FROM grade_counter")
        grade_counter = [dict(r) for r in cur.fetchall()]

        cur.execute("SELECT * FROM snapshots")
        snapshots = [dict(r) for r in cur.fetchall()]

        cur.execute("SELECT * FROM counter_snapshots")
        counter_snapshots = [dict(r) for r in cur.fetchall()]

    return {
        "exported_at": time.strftime("%Y-%m-%d %H:%M:%S"),
        "tables": {
            "users": users,
            "grade_counter": grade_counter,
            "snapshots": snapshots,
            "counter_snapshots": counter_snapshots,
        }
    }

@bot.message_handler(commands=["export"])
def export_all(message):
    chat_id = message.chat.id
    if ADMIN_CHAT_ID == 0 or chat_id != ADMIN_CHAT_ID:
        safe_send(chat_id, "❌ Команда доступна только администратору.")
        return

    payload = dump_all_data_to_dict()
    fname = f"backup_all_{time.strftime('%Y%m%d_%H%M%S')}.json"
    with open(fname, "w", encoding="utf-8") as f:
        json.dump(payload, f, ensure_ascii=False, indent=2)

    try:
        with open(fname, "rb") as f:
            bot.send_document(chat_id, f, caption="✅ Экспорт всех данных (JSON).")
    finally:
        try:
            os.remove(fname)
        except Exception:
            pass

@bot.message_handler(commands=["export_db"])
def export_db(message):
    chat_id = message.chat.id
    if ADMIN_CHAT_ID == 0 or chat_id != ADMIN_CHAT_ID:
        safe_send(chat_id, "❌ Команда доступна только администратору.")
        return
    try:
        with open(DB_PATH, "rb") as f:
            bot.send_document(chat_id, f, caption="✅ Экспорт базы SQLite (bot.db).")
    except Exception:
        safe_send(chat_id, "Не получилось отправить файл базы. Возможно его нет или нет доступа.")

# ================== Excel -> оценки ==================
def parse_excel_grades(file_path: str) -> List[Tuple[str, int]]:
    wb = load_workbook(file_path)
    sheet = wb.active
    items: List[Tuple[str, int]] = []

    for row in sheet.iter_rows(values_only=True):
        subject = row[0]
        if not subject or not isinstance(subject, str):
            continue
        subject = subject.strip()
        for cell in row[1:]:
            if isinstance(cell, (int, float)):
                items.append((subject, int(cell)))
    return items

def analyze_items(items: List[Tuple[str, int]]) -> Optional[Dict[str, Any]]:
    if not items:
        return None
    by_subject: Dict[str, List[int]] = {}
    for subj, grade in items:
        by_subject.setdefault(subj, []).append(grade)
    averages = {s: sum(vals) / len(vals) for s, vals in by_subject.items()}
    overall = sum(averages.values()) / len(averages)
    best = max(averages, key=averages.get)
    worst = min(averages, key=averages.get)
    return {"overall": overall, "best": best, "worst": worst, "averages": averages}

def make_counter(items: List[Tuple[str, int]]) -> Counter:
    c = Counter()
    for subj, grade in items:
        c[f"{subj}{SEP}{grade}"] += 1
    return c

def parse_counter_key(key: str) -> Tuple[str, int]:
    subj, grade = key.split(SEP, 1)
    return subj, int(grade)

def diff_new_grades(old: Counter, new: Counter) -> List[Tuple[str, int, int]]:
    added = []
    for key, new_count in new.items():
        old_count = old.get(key, 0)
        if new_count > old_count:
            subj, grade = parse_counter_key(key)
            added.append((subj, grade, new_count - old_count))
    added.sort(key=lambda x: (x[0], x[1]))
    return added

# ================== UI ==================
def menu_kb() -> types.InlineKeyboardMarkup:
    kb = types.InlineKeyboardMarkup(row_width=2)
    kb.add(
        types.InlineKeyboardButton("📊 Общий анализ", callback_data="summary"),
        types.InlineKeyboardButton("📚 Подробный отчёт", callback_data="details"),
    )
    kb.add(
        types.InlineKeyboardButton("📈 Динамика", callback_data="trend"),
        types.InlineKeyboardButton("🔄 Обновить данные", callback_data="refresh"),
    )
    kb.add(types.InlineKeyboardButton("↩️ Отменить последнюю загрузку", callback_data="undo"))
    kb.add(types.InlineKeyboardButton("⏰ Напоминания", callback_data="reminders"))
    return kb

def reminders_kb(enabled: bool) -> types.InlineKeyboardMarkup:
    kb = types.InlineKeyboardMarkup(row_width=2)
    toggle_text = "⛔ Выкл напоминания" if enabled else "✅ Вкл напоминания"
    kb.add(types.InlineKeyboardButton(toggle_text, callback_data="rem_toggle"))
    kb.add(
        types.InlineKeyboardButton("08:00", callback_data="time_08:00"),
        types.InlineKeyboardButton("12:00", callback_data="time_12:00"),
        types.InlineKeyboardButton("18:00", callback_data="time_18:00"),
        types.InlineKeyboardButton("21:00", callback_data="time_21:00"),
    )
    kb.add(types.InlineKeyboardButton("✍️ Ввести своё время", callback_data="time_custom"))
    kb.add(types.InlineKeyboardButton("⬅️ Назад", callback_data="back"))
    return kb

def subjects_kb(subjects: List[str], page: int = 0, per_page: int = 8) -> types.InlineKeyboardMarkup:
    kb = types.InlineKeyboardMarkup(row_width=2)
    subjects_sorted = sorted(subjects)
    start = page * per_page
    chunk = subjects_sorted[start:start + per_page]

    for s in chunk:
        kb.add(types.InlineKeyboardButton(s, callback_data=f"subj:{s}"))

    nav = []
    if page > 0:
        nav.append(types.InlineKeyboardButton("⬅️", callback_data=f"subjpage:{page-1}"))
    if start + per_page < len(subjects_sorted):
        nav.append(types.InlineKeyboardButton("➡️", callback_data=f"subjpage:{page+1}"))
    if nav:
        kb.row(*nav)

    kb.add(types.InlineKeyboardButton("⬅️ Назад", callback_data="trend"))
    return kb

# ================== Напоминания ==================
def reminder_job(chat_id: int):
    safe_send(chat_id, "⏰ Пора обновить оценки: отправь свежий Excel-файл (.xlsx).")

def schedule_user_reminder(chat_id: int, hhmm: str):
    old_job_id = scheduled_jobs.get(chat_id)
    if old_job_id:
        try:
            scheduler.remove_job(old_job_id)
        except Exception:
            pass

    hour, minute = hhmm.split(":")
    job_id = f"rem_{chat_id}"

    scheduler.add_job(
        reminder_job,
        trigger=CronTrigger(hour=int(hour), minute=int(minute)),
        args=[chat_id],
        id=job_id,
        replace_existing=True,
    )
    scheduled_jobs[chat_id] = job_id

def unschedule_user_reminder(chat_id: int):
    job_id = scheduled_jobs.get(chat_id)
    if job_id:
        try:
            scheduler.remove_job(job_id)
        except Exception:
            pass
        scheduled_jobs.pop(chat_id, None)

def restore_jobs_from_db():
    with db_lock:
        cur = CON.cursor()
        cur.execute("SELECT chat_id, reminder_time FROM users WHERE reminder_enabled=1 AND reminder_time IS NOT NULL")
        rows = cur.fetchall()
    for r in rows:
        schedule_user_reminder(int(r["chat_id"]), r["reminder_time"])

restore_jobs_from_db()

# ================== Команды ==================
@bot.message_handler(commands=["start"])
def start_cmd(message):
    ensure_user(message.chat.id)
    safe_send(
        message.chat.id,
        "Привет! 👋\n"
        "Отправь Excel (.xlsx) с оценками — я сделаю анализ.\n"
        "Данные хранятся в SQLite, есть откат последней загрузки и экспорт.",
        reply_markup=menu_kb()
    )

# ================== Приём файла ==================
@bot.message_handler(content_types=["document"])
def on_document(message):
    file_name = message.document.file_name or ""
    if not file_name.lower().endswith(".xlsx"):
        safe_send(message.chat.id, "Нужен файл формата .xlsx 🙂", reply_markup=menu_kb())
        return

    ensure_user(message.chat.id)

    try:
        file_info = bot.get_file(message.document.file_id)
        raw = bot.download_file(file_info.file_path)
    except Exception:
        safe_send(message.chat.id, "Не получилось скачать файл. Попробуй ещё раз 🙂", reply_markup=menu_kb())
        return

    tmp_name = f"{message.from_user.id}_{int(time.time())}.xlsx"
    with open(tmp_name, "wb") as f:
        f.write(raw)

    try:
        items = parse_excel_grades(tmp_name)
        rep = analyze_items(items)
        if not rep:
            safe_send(message.chat.id, "Не нашёл оценок в файле 😔", reply_markup=menu_kb())
            return

        old_counter = get_counter(message.chat.id)
        new_counter = make_counter(items)
        added = diff_new_grades(old_counter, new_counter)

        # сохранить текущее состояние
        set_counter(message.chat.id, new_counter)
        set_user_fields(
            message.chat.id,
            last_overall=float(rep["overall"]),
            last_averages_json=json.dumps(rep["averages"], ensure_ascii=False),
        )

        # snapshot + counter snapshot (для undo)
        stamp = time.strftime("%Y-%m-%d %H:%M")
        snapshot_id = add_snapshot(message.chat.id, stamp, rep["overall"], rep["averages"])
        save_counter_snapshot(snapshot_id, new_counter)

        msg = "✅ Файл обработан.\n"
        if added:
            msg += "\n🔔 Найдены новые оценки:\n"
            lines = []
            for subj, grade, cnt in added[:30]:
                suffix = f" x{cnt}" if cnt > 1 else ""
                lines.append(f"• {subj}: {grade}{suffix}")
            msg += "\n".join(lines)
            if len(added) > 30:
                msg += f"\n…и ещё {len(added) - 30}"
        else:
            msg += "\nНовых оценок не обнаружено."

        safe_send(message.chat.id, msg, reply_markup=menu_kb())

    finally:
        try:
            os.remove(tmp_name)
        except Exception:
            pass

# ================== Callback кнопок ==================
@bot.callback_query_handler(func=lambda call: True)
def on_callback(call):
    chat_id = call.message.chat.id
    ensure_user(chat_id)
    row = get_user_row(chat_id)

    if call.data == "summary":
        overall = get_last_overall(chat_id)
        averages = get_last_averages(chat_id)
        if overall is None or not averages:
            bot.answer_callback_query(call.id, "Сначала отправь Excel 🙂")
            return
        best = max(averages, key=averages.get)
        worst = min(averages, key=averages.get)
        safe_send(chat_id,
                  f"📊 Средний балл: {overall:.2f}\n🏆 Лучший предмет: {best}\n⚠ Самый слабый предмет: {worst}",
                  reply_markup=menu_kb())
        bot.answer_callback_query(call.id)
        return

    if call.data == "details":
        averages = get_last_averages(chat_id)
        if not averages:
            bot.answer_callback_query(call.id, "Сначала отправь Excel 🙂")
            return
        lines = ["📚 Отчёт по предметам:"]
        for subj, avg in sorted(averages.items(), key=lambda x: x[0]):
            lines.append(f"• {subj}: {avg:.2f}")
        safe_send(chat_id, "\n".join(lines), reply_markup=menu_kb())
        bot.answer_callback_query(call.id)
        return

    if call.data == "refresh":
        safe_send(chat_id, "🔄 Ок! Пришли новый Excel-файл (.xlsx).", reply_markup=menu_kb())
        bot.answer_callback_query(call.id)
        return

    if call.data == "trend":
        hist = get_history(chat_id, limit=10)
        if len(hist) < 2:
            bot.answer_callback_query(call.id, "Нужно минимум 2 выгрузки Excel 🙂")
            return

        lines = ["📈 Динамика среднего балла (последние 10):"]
        for h in hist:
            lines.append(f"• {h['ts']}: {h['overall']:.2f}")

        delta = hist[-1]["overall"] - hist[-2]["overall"]
        if delta > 0:
            lines.append(f"\n✅ Стало лучше на +{delta:.2f}")
        elif delta < 0:
            lines.append(f"\n⚠️ Стало хуже на {delta:.2f}")
        else:
            lines.append("\n➖ Без изменений")

        averages = get_last_averages(chat_id)
        if averages:
            lines.append("\nВыбери предмет для динамики 👇")
            safe_send(chat_id, "\n".join(lines), reply_markup=subjects_kb(list(averages.keys()), page=0))
        else:
            safe_send(chat_id, "\n".join(lines), reply_markup=menu_kb())

        bot.answer_callback_query(call.id)
        return

    if call.data.startswith("subjpage:"):
        averages = get_last_averages(chat_id)
        if not averages:
            bot.answer_callback_query(call.id, "Нет данных. Сначала отправь Excel 🙂")
            return
        page = int(call.data.split(":", 1)[1])
        safe_send(chat_id, "Выбери предмет:", reply_markup=subjects_kb(list(averages.keys()), page=page))
        bot.answer_callback_query(call.id)
        return

    if call.data.startswith("subj:"):
        subject = call.data.split(":", 1)[1]
        hist = get_history(chat_id, limit=10)
        if len(hist) < 2:
            bot.answer_callback_query(call.id, "Нужно минимум 2 выгрузки Excel 🙂")
            return

        points = []
        for h in hist:
            av = h.get("averages", {}).get(subject)
            if av is not None:
                points.append((h["ts"], float(av)))

        if len(points) < 2:
            safe_send(chat_id, f"По предмету «{subject}» мало данных (нужно 2 выгрузки).", reply_markup=menu_kb())
            bot.answer_callback_query(call.id)
            return

        lines = [f"📘 Динамика по предмету: {subject} (последние 10)"]
        for ts, av in points:
            lines.append(f"• {ts}: {av:.2f}")

        delta = points[-1][1] - points[-2][1]
        if delta > 0:
            lines.append(f"\n✅ Улучшение: +{delta:.2f}")
        elif delta < 0:
            lines.append(f"\n⚠️ Ухудшение: {delta:.2f}")
        else:
            lines.append("\n➖ Без изменений")

        safe_send(chat_id, "\n".join(lines), reply_markup=menu_kb())
        bot.answer_callback_query(call.id)
        return

    if call.data == "undo":
        last_id = get_latest_snapshot_id(chat_id)
        if not last_id:
            bot.answer_callback_query(call.id, "Нечего отменять 🙂")
            return

        with db_lock:
            cur = CON.cursor()
            cur.execute("DELETE FROM snapshots WHERE id=?", (last_id,))
            CON.commit()

        prev_id = get_latest_snapshot_id(chat_id)
        if not prev_id:
            set_user_fields(chat_id, last_overall=None, last_averages_json=None)
            set_counter(chat_id, Counter())
            safe_send(chat_id, "↩️ Откатил. История пуста, данные очищены.", reply_markup=menu_kb())
            bot.answer_callback_query(call.id)
            return

        snap = get_snapshot_data(prev_id)
        prev_counter = get_counter_by_snapshot(prev_id)
        if not snap:
            bot.answer_callback_query(call.id, "Не смог восстановить данные 😕")
            return

        set_user_fields(
            chat_id,
            last_overall=float(snap["overall"]),
            last_averages_json=json.dumps(snap["averages"], ensure_ascii=False)
        )
        set_counter(chat_id, prev_counter)
        safe_send(chat_id, f"↩️ Откатил к выгрузке {snap['ts']}.", reply_markup=menu_kb())
        bot.answer_callback_query(call.id)
        return

    if call.data == "reminders":
        enabled = bool(row["reminder_enabled"])
        t = row["reminder_time"] or "не задано"
        text = (
            "⏰ Напоминания\n"
            f"Статус: {'включены ✅' if enabled else 'выключены ⛔'}\n"
            f"Время: {t}\n\n"
            "Выбери время кнопками или введи своё."
        )
        safe_send(chat_id, text, reply_markup=reminders_kb(enabled))
        bot.answer_callback_query(call.id)
        return

    if call.data == "rem_toggle":
        enabled = not bool(row["reminder_enabled"])
        set_user_fields(chat_id, reminder_enabled=1 if enabled else 0)

        if not enabled:
            unschedule_user_reminder(chat_id)
            safe_send(chat_id, "⛔ Напоминания выключены.", reply_markup=reminders_kb(False))
            bot.answer_callback_query(call.id)
            return

        row2 = get_user_row(chat_id)
        if not row2["reminder_time"]:
            safe_send(chat_id, "✅ Включил! Теперь выбери время 👇", reply_markup=reminders_kb(True))
            bot.answer_callback_query(call.id)
            return

        schedule_user_reminder(chat_id, row2["reminder_time"])
        safe_send(chat_id, f"✅ Напоминания включены ({row2['reminder_time']}).", reply_markup=reminders_kb(True))
        bot.answer_callback_query(call.id)
        return

    if call.data.startswith("time_"):
        hhmm = call.data.replace("time_", "")
        set_user_fields(chat_id, reminder_time=hhmm)
        row2 = get_user_row(chat_id)
        if row2["reminder_enabled"]:
            schedule_user_reminder(chat_id, hhmm)
        safe_send(chat_id, f"✅ Время установлено: {hhmm}", reply_markup=reminders_kb(bool(row2["reminder_enabled"])))
        bot.answer_callback_query(call.id)
        return

    if call.data == "time_custom":
        set_user_fields(chat_id, awaiting_time=1)
        safe_send(chat_id, "Напиши время в формате HH:MM (например 18:30).")
        bot.answer_callback_query(call.id)
        return

    if call.data == "back":
        safe_send(chat_id, "💬Меню:", reply_markup=menu_kb())
        bot.answer_callback_query(call.id)
        return

    bot.answer_callback_query(call.id)

# ================== Ввод своего времени ==================
@bot.message_handler(content_types=["text"])
def on_text(message):
    chat_id = message.chat.id
    ensure_user(chat_id)
    row = get_user_row(chat_id)

    if row["awaiting_time"]:
        raw = (message.text or "").strip()
        try:
            hh_s, mm_s = raw.split(":")
            hh_i = int(hh_s)
            mm_i = int(mm_s)
            if not (0 <= hh_i <= 23 and 0 <= mm_i <= 59):
                raise ValueError
        except Exception:
            set_user_fields(chat_id, awaiting_time=0)
            safe_send(chat_id, "❌ Неправильный формат. Пример: 18:30")
            return

        hhmm = f"{hh_i:02d}:{mm_i:02d}"
        set_user_fields(chat_id, awaiting_time=0, reminder_time=hhmm)

        row2 = get_user_row(chat_id)
        if row2["reminder_enabled"]:
            schedule_user_reminder(chat_id, hhmm)

        safe_send(chat_id, f"✅ Время установлено: {hhmm}", reply_markup=menu_kb())
        return

    safe_send(chat_id, "Выбери действие кнопками 👇", reply_markup=menu_kb())

# ================== авто-перезапуск polling ==================
def run_polling_forever():
    while True:
        try:
            # skip_pending=True — чтобы после перезапуска не разгребать старые апдейты
            bot.infinity_polling(timeout=20, long_polling_timeout=20, skip_pending=True)
        except Exception as e:
            # Если polling упал — подождать и подняться снова
            print(f"[polling] crashed: {e}")
            time.sleep(5)

def run_bot():
    print("Telegram bot started")
    bot.infinity_polling()

threading.Thread(target=run_bot, daemon=True).start()

print("Flask started")

port = int(os.environ.get("PORT", 10000))
app.run(host="0.0.0.0", port=port)